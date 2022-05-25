"""
Autor: Carlos Raùl Lòpez Lòpez
Email: solrac.gt@gmail.com
Fecha: 2022-05-24

Funciòn: 
    1.- Todos los archivos y carpetas deben estar en la carpeta "GestionExcel"
    1.- Lee la carpeta "buzon", separa los archivos y directorios
    2.- Los archivos que no son "xlsx" se envian a la carpeta "NoAplica"
    3.- Los libros de Excel "xlsx" se leen uno por uno y las carpetas que contengan
        se copian al libro consolidado.xlsx en la carpeta "Procesado"
    4.- Despues de copiados se mueven a la carpeta "Procesado"

Premisas: Ver en el libro consolidado.xlsx

"""
# ------------------------------------------------------------------
#          Importar Librerias
# ------------------------------------------------------------------
import openpyxl
import os
import shutil

# ------------------------------------------------------------------
#          Variables globales
# ------------------------------------------------------------------
excelFile  = []
otrosFile  = []
subCarpeta = []

path_buzon = '/GestionExcel/buzon/'
path_excel = '/GestionExcel/Procesado/'
path_otros = '/GestionExcel/NoAplica/'

# ------------------------------------------------------------------
# Separa los archivos en arreglos
# ------------------------------------------------------------------
def separa(archivo) :
	if (archivo.find('.') != -1):
		resto = archivo.find('.') - len((archivo))
		extension = archivo[resto:]
		if extension == '.xlsx' : 
			excelFile.append(archivo)
		else: 
			otrosFile.append(archivo)		
	else:
		subCarpeta.append(archivo)
        
# ------------------------------------------------------------------
# Imprime los totales de los archivos en los arreglos generados
# ------------------------------------------------------------------
def resultados():
    print('----------------------------------------')
    print('Total de archivos procesados en el buzon')
    print('----------------------------------------')
    print('Archivos Excel xlsx: ', len(excelFile))
    print(excelFile)
    print('--------------------')
    print('Otros archivos: ', len(otrosFile))
    print(otrosFile)
    print('--------------------')
    print('Carpetas: ', len(subCarpeta))
    print(subCarpeta)
    print('--------------------')

# ------------------------------------------------------------------
# Listar ficheros dentro de la carpeta con scandir
# ------------------------------------------------------------------
def lee_buzon() :
	with os.scandir(path_buzon) as ficheros :
		for fichero in ficheros :
			separa(fichero.name)

# ------------------------------------------------------------------
# Mueve los archivos que no sean xlsx a la carpeta NoAplica
# ------------------------------------------------------------------  
def excluir_archivos():
    for i in range(len(otrosFile)) :
        archivo1 = path_buzon + otrosFile[i]
        archivo2 = path_otros + otrosFile[i]

        # Copiar archivos
        #shutil.copyfile(archivo1, archivo2)

        #Mover archivos
        shutil.move(archivo1, archivo2)

# ------------------------------------------------------------------
# Consolidar hojas de libros nuevos encontrados en el buzon
# ------------------------------------------------------------------  
def consolidar():
    # Define el destino
    libro_destino = openpyxl.load_workbook(path_excel + 'consolidado.xlsx')
    nombre_hojas_d = libro_destino.sheetnames
    numero_hojas_d = len(nombre_hojas_d)

    for l in range(len(excelFile)) :
        libro = excelFile[l]

        # Define el libro original o fuente 
        libro_origen  = openpyxl.load_workbook(path_buzon + libro)
        nombre_hojas_o = libro_origen.sheetnames
        numero_hojas_o = len(nombre_hojas_o)

        # Lectura de las hojas
        # --------------------
        for i in range(numero_hojas_o) :
            hoja = nombre_hojas_o[i]

            #Verificando indice de hojas
            #print('i:', i, 'Hoja:', hoja)

            actual = libro_origen[hoja]
            libro_origen.active = actual
            max_fil = actual.max_row
            max_col = actual.max_column

            # Se crea unahoja con el mismo nombre en el libro de destino
            hoja_destino = libro_destino.create_sheet(hoja)
            nueva = libro_destino[hoja]
            libro_destino.active = nueva

            # Traslado de las celdas de cada hoja de cada libro
            for x in range(1, max_fil + 1): 
                for y in range(1, max_col + 1) :
                    cell_ori = actual.cell(row = x, column = y) # con i se selecciona la fila
                    # Ingresa valores en la celda de destino
                    cell_des = nueva.cell(row=x, column=y)
                    cell_des.value = cell_ori.value

        archivo1 = path_buzon + libro
        archivo2 = path_excel + libro
        shutil.move(archivo1, archivo2)

    libro_destino.save(path_excel  + 'consolidado.xlsx')

# ------------------------------------------------------------------
#          Ejecuciòn
# ------------------------------------------------------------------
lee_buzon()
resultados()
excluir_archivos()
consolidar()
